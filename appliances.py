from app import db
from pathlib import Path
from openpyxl import load_workbook
import openpyxl.styles
from models import Appliance, User, UserSearch, Utility
import PyPDF2
from bs4 import BeautifulSoup
import requests

db.drop_all()
db.create_all() 

#open xlsx files with appliances/wattage and average electricity rates
xlsx_file1 = Path('static', '2.1.16.xlsx')
xlsx_file2 = Path('static', 'Table_5_06_A.xlsx')

appliance_obj = openpyxl.load_workbook(xlsx_file1)
utility_obj = openpyxl.load_workbook(xlsx_file2)

wsheet_app = appliance_obj.active
wsheet_util = utility_obj.active
canadian_util_rates = requests.get("https://www.energyhub.org/electricity-prices/")
canadian_util_rate_html = BeautifulSoup(canadian_util_rates.text, 'html.parser')

## make list of appliance categories, from column A of 2.1.16.xlsx
# appliance_cats_cleaned = ['Kitchen', 'Lighting', 'Bedroom and Bathroom'...]

appliance_cats = [cell.value for cell in wsheet_app["A"] if cell.font.bold]
appliance_cats_cleaned = appliance_cats[3:]

## also extract the rows where the category titles are found, to get the correct rows for the appliances
# row_range = [(9, 14), (16, 19), (21, 22),...]
appliance_cats_rows = [cell.row for cell in wsheet_app["A"] if cell.font.bold][3:]
# append end range to rows
appliance_cats_rows.append(55)

row_range = []
for r in range(len(appliance_cats_rows)-1):
    ran = (appliance_cats_rows[r] + 1, appliance_cats_rows[r + 1] - 1)
    row_range.append(ran)


zipped = zip(appliance_cats_cleaned, row_range)
appliance_dict = dict(zipped) #{'Kitchen': (9, 14), 'Lighting': (16, 19, ...}


## make list of appliance instances to add to seed the database
appliances = []
for (cat, ra) in appliance_dict.items():
    for val in wsheet_app.iter_rows(min_row=ra[0], max_row =ra[1], values_only=True):
        appliance = Appliance(name=val[0],
                            watts=val[5],
                            category=cat)
        appliances.append(appliance)

# remove appliance from list if there is no wattage 
cleaned_appliances = [appl for appl in appliances if appl.watts != None]

#remove extra water heater entry
cleaned_appliances.pop(26)
## Supplement missing wattage information for washer, dryer, refrigerator, etc

# pdfFileObj = open('static/1min_data.pdf', 'rb')
# pdfReader = PyPDF2.PdfFileReader(pdfFileObj)

# pageObj = pdfReader.getPage(0)
# pageObj.extractText()

#Appliances wanted, wattage on line 14 of 1min_data.pdf
supplemental_appl = {
    'Washer (Normal Wash)': [66, 'Laundry Room'],
    'Dryer (Auto-regular)': [2873, 'Laundry Room'],
    'Wall AC': [24, 'Heating and Cooling'], 
    'Central AC': [2023, 'Heating and Cooling'], 
    'Range Oven': [2795, 'Kitchen'],
    'Dishwasher': [1172, 'Kitchen'],
    'Refrigerator': [366, 'Kitchen']
}

for name in supplemental_appl.keys():
    appliance = Appliance(name=name,
                            watts=supplemental_appl[name][0],
                            category=supplemental_appl[name][1])
    cleaned_appliances.append(appliance)

#Clean utility rate data from US
utility_rates = []
for val in wsheet_util.iter_rows(min_row=5, max_row=66, values_only=True):
    rate = Utility(location=val[0],
                    rate=val[1])
    utility_rates.append(rate)

#Clean utility rate data from Canada
canada_list = [td.get_text() for td in canadian_util_rate_html.find_all('td')]
#strip '¢/kWh' from rates
canada_list_cleaned = [rate.strip('¢/kWh') for rate in canada_list]

#append Utility instances to utility_rates
for val in range(0, len(canada_list_cleaned), 2):
    rate = Utility(location = canada_list_cleaned[val],
                    rate = canada_list_cleaned[val + 1])
    utility_rates.append(rate)                                                                                                                                                                                                                                                 

#add cell phone and tablet appliances
cell_phone = Appliance(name="Cell Phone Charger", watts=5, category="Home Electronics")
tablet = Appliance(name="Tablet Charger", watts=12, category="Home Electronics")

cleaned_appliances.append(cell_phone)
cleaned_appliances.append(tablet)

db.session.add_all([*utility_rates])
db.session.add_all([*cleaned_appliances])
db.session.commit()